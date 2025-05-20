from flask import abort, flash, redirect, render_template, request, url_for, session, jsonify
from controllers.car_controller import CarController
from controllers.department_controller import DepartmentController
from controllers.internship_controller import InternshipController
from controllers.user_controller import UserController
from controllers.teacher_controller import TeacherController
from controllers.student_controller import StudentController
from models.department import Department
from models.intern_type import InternType
from models.teacher import Teacher
from models.student import Student
from models.internship import Internship
import os
from utils.utils import role_required
#from models.teacher import Teacher


def init_routes(app):
    car_controller = CarController()
    department_controller = DepartmentController()
    internship_controller = InternshipController()
    user_controller = UserController()
    teacher_controller = TeacherController()
    student_controller = StudentController()

    # =============================== Car Routes ===============================
    @app.route('/car')
    @role_required('admin', 'car')
    def car_list():
        return car_controller.list_cars()

    @app.route('/car/add', methods=['GET', 'POST'])
    @role_required('admin', 'car')
    def add_car():
        if request.method == 'POST':
            return car_controller.add_car(request)
        return car_controller.show_add_car_form()

    @app.route('/car/edit/<int:car_id>', methods=['GET', 'POST'])
    @role_required('admin', 'car')
    def edit_car(car_id):
        if request.method == 'POST':
            return car_controller.edit_car(car_id, request)
        return car_controller.show_edit_car_form(car_id)

    @app.route('/car/delete/<int:car_id>', methods=['POST'])
    @role_required('admin', 'car')
    def delete_car(car_id):
        return car_controller.delete_car(car_id)

    @app.route('/car/<int:car_id>')
    def car_details(car_id):
        return car_controller.show_car_details(car_id)

    @app.route('/car/export/xlsx')
    def export_cars_xlsx_route():
        from models.car import Car
        from utils.pdf_utils import export_cars_xlsx
        cars = Car.get_all()
        xlsx_io = export_cars_xlsx(cars)
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='cars.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/car/export/pdf')
    def export_cars_pdf_route():
        from models.car import Car
        from utils.pdf_utils import export_cars_pdf
        cars = Car.get_all()
        pdf_io = export_cars_pdf(cars)
        from flask import send_file
        return send_file(pdf_io, as_attachment=True, download_name='cars.pdf', mimetype='application/pdf')

    # =============================== Department Routes ===============================
    @app.route('/department')
    def department_list():
        return department_controller.list_departments()

    @app.route('/department/add', methods=['GET', 'POST'])
    def add_department():
        if request.method == 'POST':
            return department_controller.add_department(request)
        return department_controller.show_add_department_form()

    @app.route('/department/edit/<int:department_id>', methods=['GET', 'POST'])
    def edit_department(department_id):
        if request.method == 'POST':
            return department_controller.edit_department(department_id, request)
        return department_controller.show_edit_department_form(department_id)

    @app.route('/department/delete/<int:department_id>', methods=['POST'])
    def delete_department(department_id):
        # Call the controller's delete_department method
        return department_controller.delete_department(department_id)

    @app.route('/department/export/pdf')
    def export_departments_pdf_route():
        from models.department import Department
        from utils.pdf_utils import export_departments_pdf
        departments = Department.get_all()
        pdf_io = export_departments_pdf(departments)
        from flask import send_file
        return send_file(pdf_io, as_attachment=True, download_name='departments.pdf', mimetype='application/pdf')

    @app.route('/department/export/xlsx')
    def export_departments_xlsx_route():
        from models.department import Department
        from utils.pdf_utils import export_departments_xlsx
        departments = Department.get_all()
        xlsx_io = export_departments_xlsx(departments)
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='departments.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/department/sample/xlsx')
    def sample_department_xlsx_route():
        from utils.pdf_utils import sample_department_xlsx
        xlsx_io = sample_department_xlsx()
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='sample_department_import.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/department/import/xlsx', methods=['GET', 'POST'])
    def import_departments_xlsx():
        from flask import request, redirect, url_for, flash, render_template
        from utils.pdf_utils import import_departments_from_xlsx
        from models.department import Department
        if request.method == 'POST':
            file = request.files.get('file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Veuillez sélectionner un fichier XLSX valide.', 'danger')
                return redirect(request.url)
            count, errors = import_departments_from_xlsx(file)
            if errors:
                flash(f"Import partiel: {count} départements ajoutés. Erreurs: {errors}", 'warning')
            else:
                flash(f"{count} départements importés avec succès!", 'success')
            return redirect(url_for('department_list'))
        return render_template('department/import_xlsx.html')

    # =============================== Teacher Routes ===============================
    @app.route('/teacher')
    @role_required('admin', 'teacher')
    def teacher_list():
        return teacher_controller.list_teachers()

    @app.route('/teacher/add', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def add_teacher():
        if request.method == 'POST':
            return teacher_controller.add_teacher(request)
        return teacher_controller.show_add_teacher_form()

    @app.route('/teacher/<int:teacher_id>')
    @role_required('admin', 'teacher')
    def teacher_details(teacher_id):
        return teacher_controller.show_teacher_details(teacher_id)

    @app.route('/teacher/edit/<int:teacher_id>', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def edit_teacher(teacher_id):
        if request.method == 'POST':
            return teacher_controller.edit_teacher(teacher_id, request)
        teacher = Teacher.get_by_id(teacher_id)
        departments = Department.get_all()  # Use plural 'departments'
        print("Route list of departments", departments)  # Debugging output
        return render_template('teacher/edit.html', teacher=teacher, departments=departments)


    @app.route('/teacher/delete/<int:teacher_id>', methods=['POST'])
    @role_required('admin')
    def delete_teacher(teacher_id):
        return teacher_controller.delete_teacher(teacher_id)
    
    @app.route('/teacher/export/pdf')
    def export_teachers_pdf_route():
        from models.teacher import Teacher
        from utils.pdf_utils import export_teachers_pdf
        teachers = Teacher.get_all()
        pdf_io = export_teachers_pdf(teachers)
        from flask import send_file
        return send_file(pdf_io, as_attachment=True, download_name='teachers.pdf', mimetype='application/pdf')

    @app.route('/teacher/export/xlsx')
    def export_teachers_xlsx_route():
        from models.teacher import Teacher
        from utils.pdf_utils import export_teachers_xlsx
        teachers = Teacher.get_all()
        xlsx_io = export_teachers_xlsx(teachers)
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='teachers.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/teacher/import/xlsx', methods=['GET', 'POST'])
    def import_teachers_xlsx():
        from flask import request, redirect, url_for, flash, render_template
        from utils.pdf_utils import import_teachers_from_xlsx
        from models.teacher import Teacher
        if request.method == 'POST':
            file = request.files.get('file')
            if not file or not file.filename.endswith('.xlsx'):
                flash('Veuillez sélectionner un fichier XLSX valide.', 'danger')
                return redirect(request.url)
            count, errors = import_teachers_from_xlsx(file)
            if errors:
                flash(f"Import partiel: {count} enseignants ajoutés. Erreurs: {errors}", 'warning')
            else:
                flash(f"{count} enseignants importés avec succès!", 'success')
            return redirect(url_for('teacher_list'))
        return render_template('teacher/import_xlsx.html')

    @app.route('/teacher/sample/xlsx')
    def sample_teacher_xlsx_route():
        from utils.pdf_utils import sample_teacher_xlsx
        xlsx_io = sample_teacher_xlsx()
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='sample_teacher_import.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # =============================== Internship Routes ===============================
    @app.route('/internship')
    @role_required('admin', 'teacher', 'car')
    def internship_list():
        return internship_controller.list_internships()

    @app.route('/internship/add', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def add_internship():
        if request.method == 'POST':
            return internship_controller.add_internship(request)
        # Fetch intern types from the database
        intern_types = InternType.get_all() 
        return internship_controller.show_add_internship_form(intern_types)
    
    @app.route('/internship/edit/<int:internship_id>', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def edit_internship(internship_id):
        if request.method == 'POST':
            return internship_controller.edit_internship(internship_id, request)
        # Fetch intern types from the database
        intern_types = InternType.get_all()  
        return internship_controller.show_edit_internship_form(internship_id)


    @app.route('/internship/delete/<int:internship_id>', methods=['POST'])
    @role_required('admin')
    def delete_internship(internship_id):
        return internship_controller.delete_internship(internship_id)

    @app.route('/internship/cancel/<int:internship_id>', methods=['POST'])
    @role_required('admin', 'teacher')
    def cancel_internship(internship_id):
        return internship_controller.cancel_internship(internship_id)

    @app.route('/internship/mark_done/<int:internship_id>', methods=['POST'])
    @role_required('admin', 'teacher')
    def mark_as_done(internship_id):
        return internship_controller.mark_as_done(internship_id)

    @app.route('/internship/details/<int:internship_id>')
    @role_required('admin', 'teacher', 'car')
    def internship_details(internship_id):
        return internship_controller.show_internship_details(internship_id)

    @app.route('/internship/export/xlsx')
    def export_internships_xlsx_route():
        from models.internship import Internship
        from utils.pdf_utils import export_internships_xlsx
        internships = Internship.get_all()
        xlsx_io = export_internships_xlsx(internships)
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='internships.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/internship/export/pdf')
    def export_internships_pdf_route():
        from models.internship import Internship
        from utils.pdf_utils import export_internships_pdf
        internships = Internship.get_all()
        pdf_io = export_internships_pdf(internships)
        from flask import send_file
        return send_file(pdf_io, as_attachment=True, download_name='internships.pdf', mimetype='application/pdf')
    
    # =============================== User Routes ===============================
    @app.route('/login', methods=['GET', 'POST'])
    def login_user():
        if request.method == 'POST':
            return user_controller.login(request)
        return user_controller.show_login_form()

    @app.route('/logout')
    def logout_user():
        return user_controller.logout()
    
    @app.route('/register', methods=['GET', 'POST'])
    def register_user():
        if request.method == 'POST':
            return user_controller.register(request)
        return user_controller.show_register_form()

    @app.route('/user')
    def list_users():
        return user_controller.list_users()

    @app.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])
    @role_required('admin')
    def edit_user(user_id):
        if request.method == 'POST':
            return user_controller.edit_user(user_id, request)
        return user_controller.show_edit_user_form(user_id)
    
    @app.route('/user/update_profile', methods=['POST'])
    @role_required('admin', 'teacher', 'car')
    def update_profile():
        return user_controller.update_profile(request)
    
    @app.route('/user/delete/<int:user_id>', methods=['POST'])
    def delete_user(user_id):
        return user_controller.delete_user(user_id)
    
    # =============================== Internship Type Routes ===============================
    @app.route('/internship/types')
    @role_required('admin', 'teacher')
    def list_internship_types():
        intern_types = InternType.get_all()  # Fetch all internship types from the database
        return render_template('internship/list_internship_types.html', internship_types=intern_types)

    # Add a new internship type
    @app.route('/internship/types/add', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def add_internship_type():
        if request.method == 'POST':
            name = request.form['type_name']
            InternType.add_intern_type(name)  # Add the new type to the database
            flash('Internship Type added successfully!', 'success')
            return redirect(url_for('list_internship_types'))
        return render_template('internship/add_internship_type.html')

    # Edit an internship type
    @app.route('/internship/types/edit/<int:intern_type_id>', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def edit_internship_type(intern_type_id):
        intern_type = InternType.get_by_id(intern_type_id)  # Get the intern type by ID
        if not intern_type:
            flash('Internship Type not found!', 'danger')
            return redirect(url_for('list_internship_types'))

        if request.method == 'POST':
            name = request.form['type_name']
            InternType.update_intern_type(intern_type_id, name)  # Update the internship type
            flash('Internship Type updated successfully!', 'success')
            return redirect(url_for('list_internship_types'))

        return render_template('internship/edit_internship_type.html', intern_type=intern_type)

    # Delete an internship type
    @app.route('/internship/types/delete/<int:intern_type_id>', methods=['POST'])
    @role_required('admin')
    def delete_internship_type(intern_type_id):
        intern_type = InternType.get_by_id(intern_type_id)  # Fetch the intern type to ensure it exists
        if not intern_type:
            flash('Internship Type not found!', 'danger')
            return redirect(url_for('list_internship_types'))

        InternType.delete_intern_type(intern_type_id)  # Delete the internship type
        flash('Internship Type deleted successfully!', 'success')
        return redirect(url_for('list_internship_types'))

    @app.route('/internship/types/sample/xlsx')
    def sample_intern_type_xlsx_route():
        from utils.pdf_utils import sample_intern_type_xlsx
        xlsx_io = sample_intern_type_xlsx()
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='sample_intern_type_import.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/internship/types/import/xlsx', methods=['GET', 'POST'])
    def import_intern_types_xlsx():
        from flask import request, redirect, url_for, flash, render_template
        from utils.pdf_utils import import_intern_types_from_xlsx
        if request.method == 'POST':
            if 'file' not in request.files or not request.files['file'].filename.endswith('.xlsx'):
                flash('Veuillez sélectionner un fichier XLSX valide.', 'danger')
                return redirect(request.url)
            file = request.files['file']
            count, errors = import_intern_types_from_xlsx(file)
            if errors:
                flash(f"Import partiel: {count} types ajoutés. Erreurs: {errors}", 'warning')
            else:
                flash(f"{count} types importés avec succès!", 'success')
            return redirect(url_for('list_internship_types'))
        return render_template('internship/import_xlsx.html')

    @app.route('/internship/types/export/xlsx')
    def export_internship_types_xlsx_route():
        from models.intern_type import InternType
        from utils.pdf_utils import export_internship_types_xlsx
        intern_types = InternType.get_all()
        xlsx_io = export_internship_types_xlsx(intern_types)
        from flask import send_file
        return send_file(xlsx_io, as_attachment=True, download_name='internship_types.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # =============================== Misc/Utility Routes ===============================
    @app.route('/refresh_app', methods=['POST'])
    def refresh_app():
        if 'role' not in session or session['role'] != 'admin':
            flash('Unauthorized: Only admin can refresh the app.', 'danger')
            return redirect(url_for('index'))
        # Touch app.py to trigger Flask reloader
        app_py_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.py')
        try:
            os.utime(app_py_path, None)
            flash('App code refresh triggered!', 'info')
        except Exception as e:
            flash(f'Failed to refresh app: {e}', 'danger')
        return redirect(url_for('index'))

    @app.route('/logs')
    def logs_page():
        if 'role' not in session or session['role'] != 'admin':
            flash('Unauthorized: Only admin can access logs.', 'danger')
            return redirect(url_for('index'))
        import os, json
        log_path = os.path.join(os.path.dirname(__file__), 'logs', 'logs.json')
        logs = []
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        return render_template('logs.html', logs=logs)

    @app.route('/logs/json')
    def logs_json():
        if 'role' not in session or session['role'] != 'admin':
            flash('Unauthorized: Only admin can access logs.', 'danger')
            return redirect(url_for('index'))
        import os, json
        log_path = os.path.join(os.path.dirname(__file__), 'logs', 'logs.json')
        if not os.path.exists(log_path):
            return jsonify([])
        with open(log_path, 'r', encoding='utf-8') as f:
            logs = json.load(f)
        return jsonify(logs)

    @app.route('/logs/export/xlsx')
    def export_logs_xlsx():
        if 'role' not in session or session['role'] != 'admin':
            flash('Unauthorized: Only admin can export logs.', 'danger')
            return redirect(url_for('index'))
        import os, json
        from io import BytesIO
        from openpyxl import Workbook
        from flask import send_file
        log_path = os.path.join(os.path.dirname(__file__), 'logs', 'logs.json')
        logs = []
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Logs'
        headers = ['Date & Heure', "Nom d'utilisateur", 'Rôle', 'IP', 'Action', 'Table', 'ID Élément']
        ws.append(headers)
        for log in logs:
            ws.append([
                log.get('timestamp', '').replace('T', ' '),
                log.get('username', ''),
                log.get('role', ''),
                log.get('ip', ''),
                log.get('action', ''),
                log.get('table', ''),
                log.get('item_id', '')
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='logs.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # =============================== Student Routes ===============================
    @app.route('/student')
    def student_list():
        return student_controller.list_students()

    @app.route('/student/add', methods=['GET', 'POST'])
    def add_student():
        internship_id = request.args.get('internship_id')
        if request.method == 'POST':
            student_id = student_controller.add_student(request)
            if internship_id:
                Student.assign_internship(student_id, internship_id)
            return redirect('/student')
        return student_controller.show_add_student_form(internship_id=internship_id)

    @app.route('/internship/<int:internship_id>/add_student', methods=['GET', 'POST'])
    def add_student_to_internship(internship_id):
        from controllers.student_controller import StudentController
        student_controller = StudentController()
        if request.method == 'POST':
            return student_controller.add_student_to_internship(internship_id, request)
        # Optionally, show a form if GET is needed
        return redirect(f'/internship/details/{internship_id}')

    @app.route('/internship/choose_and_add_student', methods=['GET', 'POST'])
    def choose_internship_and_add_student():
        if request.method == 'POST':
            internship_id = request.form['internship_id']
            return redirect(f'/internship/{internship_id}/add_student')
        internships = Internship.get_all()
        return render_template('student/choose_internship.html', internships=internships)

    @app.route('/student/import/xlsx', methods=['POST'])
    def import_students_xlsx():
        from controllers.student_controller import StudentController
        import openpyxl
        student_controller = StudentController()
        file = request.files['file']
        internship_id = request.form.get('internship_id')
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            first_name, last_name, email, phone = row[:4]
            rows.append({'first_name': first_name, 'last_name': last_name, 'email': email, 'phone': phone})
        student_controller.import_students_xlsx(rows, internship_id=internship_id)
        return redirect(f'/internship/details/{internship_id}')

    @app.route('/student/sample/xlsx')
    def sample_student_xlsx():
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['first_name', 'last_name', 'email', 'phone'])
        ws.append(['John', 'Doe', 'john@example.com', '123456789'])
        ws.append(['Jane', 'Smith', '', ''])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(output, as_attachment=True, download_name='students_sample.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/student/export/pdf')
    def export_students_pdf():
        from models.student import Student
        from utils.pdf_utils import generate_students_pdf
        students = Student.get_all()
        pdf_bytes = generate_students_pdf(students)
        from flask import send_file
        import io
        return send_file(io.BytesIO(pdf_bytes), as_attachment=True, download_name='students.pdf', mimetype='application/pdf')

    @app.route('/student/export/xlsx')
    def export_students_xlsx():
        from models.student import Student
        import openpyxl
        from flask import send_file
        import io
        students = Student.get_all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['first_name', 'last_name', 'email', 'phone', 'internship', 'intern_type', 'teacher'])
        for s in students:
            ws.append([
                s['first_name'],
                s['last_name'],
                s['email'],
                s['phone'],
                f"{s['internship_start']} - {s['internship_end']}" if s['internship_start'] else '',
                s['intern_type_name'] or '',
                f"{s['teacher_first_name']} {s['teacher_last_name']}" if s['teacher_first_name'] else ''
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='students.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/internship/<int:internship_id>/students/export/pdf')
    def export_students_of_internship_pdf(internship_id):
        from models.student import Student
        from utils.pdf_utils import generate_students_pdf
        students = [s for s in Student.get_all() if s.get('internship_id') == internship_id]
        pdf_bytes = generate_students_pdf(students)
        from flask import send_file
        import io
        return send_file(io.BytesIO(pdf_bytes), as_attachment=True, download_name=f'students_internship_{internship_id}.pdf', mimetype='application/pdf')

    @app.route('/internship/<int:internship_id>/students/export/xlsx')
    def export_students_of_internship_xlsx(internship_id):
        from models.student import Student
        from openpyxl import Workbook
        from flask import send_file
        import io
        students = [s for s in Student.get_all() if s.get('internship_id') == internship_id]
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["First Name", "Last Name", "Email", "Phone"])
        for s in students:
            ws.append([
                s.get('first_name', ''),
                s.get('last_name', ''),
                s.get('email', ''),
                s.get('phone', '')
            ])
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)
        return send_file(xlsx_io, as_attachment=True, download_name=f'students_internship_{internship_id}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/student/edit/<int:student_id>', methods=['GET', 'POST'])
    @role_required('admin', 'teacher')
    def edit_student(student_id):
        from controllers.student_controller import StudentController
        return StudentController().edit_student(student_id)

    @app.route('/student/delete/<int:student_id>', methods=['POST'])
    @role_required('admin', 'teacher')
    def delete_student(student_id):
        from controllers.student_controller import StudentController
        return StudentController().delete_student(student_id)
