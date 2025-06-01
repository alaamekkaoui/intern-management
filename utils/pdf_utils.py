from fpdf import FPDF
import io
import openpyxl
from openpyxl.workbook import Workbook

def add_iav_logo(pdf):
    # Add the IAV logo centered at the top
    logo_path = 'static/images/iav.png'
    page_width = pdf.w - 2 * pdf.l_margin
    logo_width = 40
    x = (pdf.w - logo_width) / 2
    pdf.image(logo_path, x=x, y=10, w=logo_width)
    pdf.ln(30)  # Add space below the logo

def export_teachers_pdf(teachers):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des Enseignants', ln=True, align='C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Prénom', 1)
    pdf.cell(40, 10, 'Nom', 1)
    pdf.cell(60, 10, 'E-mail', 1)
    pdf.cell(30, 10, 'Téléphone', 1)
    pdf.cell(0, 10, 'Département', 1, ln=True)
    pdf.set_font('Arial', '', 12)
    for t in teachers:
        pdf.cell(40, 10, str(t.get('first_name', '')), 1)
        pdf.cell(40, 10, str(t.get('last_name', '')), 1)
        pdf.cell(60, 10, str(t.get('email', '')), 1)
        pdf.cell(30, 10, str(t.get('phone', '')), 1)
        pdf.cell(0, 10, str(t.get('department_name', '')), 1, ln=True)
    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def sample_teacher_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    # Header
    ws.append(["First Name", "Last Name", "Email", "Phone", "Department"])
    # Sample row
    ws.append(["John", "Doe", "john.doe@example.com", "1234567890", "Computer Science"])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def export_teachers_xlsx(teachers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Enseignants"
    ws.append(["Prénom", "Nom", "E-mail", "Téléphone", "Département"])
    for t in teachers:
        ws.append([
            t.get('first_name', ''),
            t.get('last_name', ''),
            t.get('email', ''),
            t.get('phone', ''),
            t.get('department_name', '')
        ])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def import_teachers_from_xlsx(file):
    from models.teacher import Teacher
    from models.department import Department
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    count = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < 5 or all(cell is None for cell in row):
            errors.append(f"Ligne {i}: Données manquantes ou ligne vide.")
            continue
        first_name, last_name, email, phone, department_name = row[:5]
        if not (first_name and last_name and email and department_name):
            errors.append(f"Ligne {i}: Données manquantes.")
            continue
        department = Department.get_by_name(department_name)
        if not department:
            errors.append(f"Ligne {i}: Département '{department_name}' introuvable.")
            continue
        success = Teacher.add_teacher(first_name, last_name, department['id'], email, phone)
        if success:
            count += 1
        else:
            errors.append(f"Ligne {i}: Erreur lors de l'ajout de l'enseignant.")
    return count, errors

def export_internships_xlsx(internships):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stages"
    ws.append(["Titre", "Enseignant", "Date de Début", "Date de Fin", "Statut", "Voiture", "Type de Stage"])
    for i in internships:
        ws.append([
            i.get('title', ''),
            f"{i.get('first_name', '')} {i.get('last_name', '')}",
            i.get('start_date', ''),
            i.get('end_date', ''),
            i.get('status', ''),
            i.get('car_model', ''),
            i.get('intern_type', '')
        ])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def export_cars_xlsx(cars):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voitures"
    ws.append(["Nom", "Modèle", "Type", "Marque", "Plaque d'Immatriculation", "Coût/Jour", "Nombre Disponible"])
    for c in cars:
        ws.append([
            c.get('name', ''),
            c.get('model', ''),
            c.get('car_type', ''),
            c.get('brand', ''),
            c.get('license_plate', ''),
            c.get('average_cost_per_day', ''),
            c.get('available_count', '')
        ])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def export_departments_xlsx(departments):
    wb = Workbook()
    ws = wb.active
    ws.title = "Départements"
    ws.append(["Nom"])
    for d in departments:
        ws.append([d.get('name', '')])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def sample_department_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Name"])
    ws.append(["Mathematics"])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def import_departments_from_xlsx(file):
    from models.department import Department
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    count = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < 1 or not row[0]:
            errors.append(f"Ligne {i}: Nom du département manquant.")
            continue
        name = row[0]
        existing = Department.get_by_name(name)
        if existing:
            errors.append(f"Ligne {i}: Département '{name}' existe déjà.")
            continue
        success = Department.add_department(name)
        if success:
            count += 1
        else:
            errors.append(f"Ligne {i}: Erreur lors de l'ajout du département.")
    return count, errors

def export_departments_pdf(departments):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des Départements', ln=True, align='C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Nom', 1, ln=True)
    pdf.set_font('Arial', '', 12)
    for d in departments:
        pdf.cell(0, 10, d.get('name', ''), 1, ln=True)
    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def export_internships_pdf(internships):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des Stages', ln=True, align='C')
    pdf.ln(10)

    # Define column widths and headers
    page_width = pdf.w - 2 * pdf.l_margin
    col_widths = [40, 48, 18, 18, 18, 28, 18] # Total width: 188. OK.
    col_headers = ['Titre', 'Enseignant', 'Début', 'Fin', 'Statut', 'Voiture', 'Type']
    line_height = 5 # Base line height for content rows

    pdf.set_font('Arial', 'B', 9) # Smaller font for headers
    # Print table headers
    for col_header, col_width in zip(col_headers, col_widths):
        pdf.cell(col_width, 10, col_header, 1, 0, 'C')
    pdf.ln() # Move to the next line after printing headers

    pdf.set_font('Arial', '', 8) # Smaller font for content

    # Print table rows
    for i in internships:
        row_data = [
            str(i.get('title', '')),
            f"{i.get('first_name', '')} {i.get('last_name', '')}",
            str(i.get('start_date', '')),
            str(i.get('end_date', '')),
            str(i.get('status', '')),
            str(i.get('car_model', '')),
            str(i.get('intern_type', ''))
        ]

        # Calculate max height for the current row based on its content and column widths
        max_row_height = line_height # Start with base height
        for col_data, col_width in zip(row_data, col_widths):
            # Temporarily set font to calculate string width accurately
            pdf.set_font('Arial', '', 8)
            text_width = pdf.get_string_width(col_data)
            # Estimate lines needed, using ceil for integer division
            lines = int(text_width / col_width) + (text_width % col_width > 0)
            lines = max(1, lines) # Ensure at least one line
            cell_calculated_height = lines * line_height
            max_row_height = max(max_row_height, cell_calculated_height)

        # Ensure a minimum row height (e.g., for empty cells or minimal content)
        # Set a reasonable minimum height if content is very short or empty
        min_row_height = line_height * 2.5 # Example: minimum height equivalent to 2.5 lines
        max_row_height = max(min_row_height, max_row_height)

        # Store the initial y position for the row
        start_y = pdf.get_y()
        start_x = pdf.get_x() # Store starting x for the row

        # Print cells for the row
        current_x = start_x
        for col_data, col_width in zip(row_data, col_widths):
            # Draw the multi-line cell content
            # Save current position before multi_cell as it moves Y
            before_multi_cell_y = pdf.get_y()
            pdf.set_xy(current_x, before_multi_cell_y) # Ensure starting at current_x, current_y
            pdf.multi_cell(col_width, line_height, col_data, border=0, align='L', fill=0)

            # Draw the cell border with the calculated max_row_height
            # After multi_cell, the cursor is below the text. Need to reset X and Y for drawing border.
            pdf.set_xy(current_x, start_y); # Reset position to the start of the cell area
            pdf.cell(col_width, max_row_height, '', border=1, ln=0, align='L', fill=0); # Draw border with max_row_height

            # Move to the start of the next column, staying at the initial row Y level
            current_x += col_width;
            pdf.set_xy(current_x, start_y); # Position for the next cell's content and border start

        # After printing all cells in the row, move to the next line based on the maximum row height
        pdf.set_y(start_y + max_row_height);

    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1') # Use latin1 for broader character support
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def export_cars_pdf(cars):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des Voitures', ln=True, align='C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(30, 10, 'Nom', 1)
    pdf.cell(25, 10, 'Modèle', 1)
    pdf.cell(25, 10, 'Type', 1)
    pdf.cell(25, 10, 'Marque', 1)
    pdf.cell(30, 10, 'Plaque', 1)
    pdf.cell(25, 10, 'Coût/Jour', 1)
    pdf.cell(0, 10, 'Disponible', 1, ln=True)
    pdf.set_font('Arial', '', 10)
    for c in cars:
        pdf.cell(30, 10, str(c.get('name', '')), 1)
        pdf.cell(25, 10, str(c.get('model', '')), 1)
        pdf.cell(25, 10, str(c.get('car_type', '')), 1)
        pdf.cell(25, 10, str(c.get('brand', '')), 1)
        pdf.cell(30, 10, str(c.get('license_plate', '')), 1)
        pdf.cell(25, 10, str(c.get('average_cost_per_day', '')), 1)
        pdf.cell(0, 10, str(c.get('available_count', '')), 1, ln=True)
    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def sample_intern_type_xlsx():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Internship Types"
    ws.append(["Name"])
    ws.append(["Stage terrain"])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def import_intern_types_from_xlsx(file):
    from models.intern_type import InternType
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    count = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < 1 or not row[0]:
            errors.append(f"Ligne {i}: Nom du type manquant.")
            continue
        name = row[0]
        # Check for duplicate
        existing = [t for t in InternType.get_all() if t['name'].strip().lower() == name.strip().lower()]
        if existing:
            errors.append(f"Ligne {i}: Type '{name}' existe déjà.")
            continue
        try:
            InternType.add_intern_type(name)
            count += 1
        except Exception as e:
            errors.append(f"Ligne {i}: Erreur lors de l'ajout du type: {e}")
    return count, errors

def export_internship_types_xlsx(internship_types):
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Internship Types"
    ws.append(["Name"])
    for t in internship_types:
        ws.append([t.get('name', '')])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def generate_students_pdf(students):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des Étudiants', ln=True, align='C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Prénom', 1)
    pdf.cell(40, 10, 'Nom', 1)
    pdf.cell(60, 10, 'E-mail', 1)
    pdf.cell(30, 10, 'Téléphone', 1)
    pdf.cell(0, 10, 'Stage', 1, ln=True)
    pdf.set_font('Arial', '', 12)
    for s in students:
        pdf.cell(40, 10, str(s.get('first_name', '')), 1)
        pdf.cell(40, 10, str(s.get('last_name', '')), 1)
        pdf.cell(60, 10, str(s.get('email', '')), 1)
        pdf.cell(30, 10, str(s.get('phone', '')), 1)
        internship = s.get('internship_title', '') if s.get('internship_title') else ''
        pdf.cell(0, 10, internship, 1, ln=True)
    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def sample_student_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["first_name", "last_name", "email", "phone", "internship_id", "teacher_id"])
    ws.append(["John", "Doe", "john.doe@email.com", "0612345678", 1, 1])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def sample_car_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cars"
    ws.append(["Name", "Model", "Type", "Brand", "License Plate", "Cost/Day", "Available Count"])
    ws.append(["Car A", "4x4", "SUV", "Toyota", "ABC-123", 100.00, 2])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def sample_internship_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Internships"
    ws.append(["Title", "Teacher ID", "Intern Type ID", "Car ID", "Start Date", "End Date", "Status", "Num Ordre Mission", "Description", "Destination", "Kilometrage"])
    ws.append(["AI Research", 1, 1, 1, "2024-07-01", "2024-08-01", "pending", "ORD-001", "Research on AI", "Rabat", 120])
    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io

def import_cars_from_xlsx(file):
    from models.car import Car
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    count = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < 7 or all(cell is None for cell in row):
            errors.append(f"Ligne {i}: Données manquantes ou ligne vide.")
            continue
        name, model, car_type, brand, license_plate, cost_per_day, available_count = row[:7]
        if not (name and model and car_type and brand and license_plate and cost_per_day is not None and available_count is not None):
            errors.append(f"Ligne {i}: Données manquantes.")
            continue
        try:
            success = Car.add_car(name, model, car_type, brand, license_plate, cost_per_day, available_count)
            if success:
                count += 1
            else:
                errors.append(f"Ligne {i}: Erreur lors de l'ajout de la voiture.")
        except Exception as e:
            errors.append(f"Ligne {i}: Erreur lors de l'ajout de la voiture: {e}")
    return count, errors

def import_internships_from_xlsx(file):
    from models.internship import Internship
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    count = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < 11 or all(cell is None for cell in row):
            errors.append(f"Ligne {i}: Données manquantes ou ligne vide.")
            continue
        title, teacher_id, intern_type_id, car_id, start_date, end_date, status, num_ordre_mission, description, destination, kilometrage = row[:11]
        if not (title and teacher_id and intern_type_id and start_date and end_date and status):
            errors.append(f"Ligne {i}: Données obligatoires manquantes.")
            continue
        try:
            success = Internship.add_internship(teacher_id, intern_type_id, car_id, start_date, end_date, num_ordre_mission, description, destination, kilometrage)
            if success:
                count += 1
            else:
                errors.append(f"Ligne {i}: Erreur lors de l'ajout du stage.")
        except Exception as e:
            errors.append(f"Ligne {i}: Erreur lors de l'ajout du stage: {e}")
    return count, errors

def export_single_internship_pdf(internship):
    pdf = FPDF()
    pdf.add_page()
    add_iav_logo(pdf)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Détails du Stage', ln=True, align='C')
    pdf.ln(10)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Titre:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.multi_cell(0, 10, str(internship.get('title', '')))

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Enseignant:', 0)
    pdf.set_font('Arial', '', 12)
    teacher_name = f"{internship.get('first_name', '')} {internship.get('last_name', '')}"
    pdf.multi_cell(0, 10, teacher_name)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Date de Début:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, str(internship.get('start_date', '')), ln=True)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Date de Fin:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, str(internship.get('end_date', '')), ln=True)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Statut:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, str(internship.get('status', '')), ln=True)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Voiture:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, str(internship.get('car_model', '')), ln=True)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(40, 10, 'Type de Stage:', 0)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, str(internship.get('intern_type', '')), ln=True)

    # Add other details if available (num_ordre_mission, description, destination, kilometrage)
    if internship.get('num_ordre_mission'):
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, 'Numéro d\'ordre de Mission:', 0)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, str(internship.get('num_ordre_mission', '')), ln=True)

    if internship.get('description'):
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, 'Description:', 0)
        pdf.set_font('Arial', '', 12)
        pdf.multi_cell(0, 10, str(internship.get('description', '')))

    if internship.get('destination'):
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, 'Destination:', 0)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, str(internship.get('destination', '')), ln=True)

    if internship.get('kilometrage'):
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, 'Kilométrage:', 0)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, str(internship.get('kilometrage', '')), ln=True)


    pdf_output = io.BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_bytes)
    pdf_output.seek(0)
    return pdf_output

def export_students_xlsx(students):
    """
    Exports a list of students to an XLSX file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    ws.append(["Prénom", "Nom", "E-mail", "Téléphone", "Stage", "Type de Stage", "Enseignant"]) # Add headers

    for s in students:
        internship_title = s.get('internship_title', '') if s.get('internship_title') else ''
        intern_type_name = s.get('intern_type_name', '') if s.get('intern_type_name') else ''
        teacher_name = f"{s.get('teacher_first_name', '')} {s.get('teacher_last_name', '')}".strip()

        ws.append([
            s.get('first_name', ''),
            s.get('last_name', ''),
            s.get('email', ''),
            s.get('phone', ''),
            internship_title,
            intern_type_name,
            teacher_name
        ])

    xlsx_io = io.BytesIO()
    wb.save(xlsx_io)
    xlsx_io.seek(0)
    return xlsx_io